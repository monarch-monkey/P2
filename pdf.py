from openpyxl import load_workbook
from os import remove,rmdir,listdir,rename
from shutil import copyfile

def getfilm(li):
    path="C:/Users/王杰/Desktop/教学/python课程导读pdf版"
    path_vedio="C:/Users/王杰/Desktop/"
    files=listdir(path)
    for y in li:
        if y[0]=="P1":  
            path_pro=path+"/P1"
        if y[0]=="P2":  
            path_pro=path+"/P2"     
        files_pdfs=listdir(path_pro)
        for g in files_pdfs:
            index=g.split(".")[0][-2::]
            if index==y[1]:
                copyfile(path_pro+"/"+g,path_vedio+"/"+g)
                try:
                    rename(path_vedio+"/"+g,path_vedio+"/"+y[2]+".pdf")
                except FileExistsError:
                    pass 
def getTableStr():
    wb = load_workbook("tablelist.xlsx")
    ws=wb.active
    i=3
    msg=[]
    s=ws["A2"].value+"\n"
    while True:
        word=ws["B"+str(i)].value
        if ws["B"+str(i)].value == None:
            break
        else:
            if word[0]=="P":    
                word1=ws["C"+str(i)].value
                if word[5]=="课":
                    msg.append([word[:2],"0"+word[4],word1]) 
                else:    
                    msg.append([word[:2],word[4:6],word1])
                word2=ws["A"+str(i)].value
                s += word2+" "+word+" "+word1+"\n"
        i+=1    
    return msg 
if __name__ == "__main__":
    getfilm(getTableStr())                
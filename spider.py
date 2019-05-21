from os import listdir, remove, rename, rmdir
from shutil import copyfile
from time import sleep

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options


def find_taday_table(date):
    # 登入当前课表界面
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(chrome_options=chrome_options)
    driver.set_window_size(1366,768)
    driver.get("https://vts.vipcode.com/")
    sleep(1)
    driver.find_element_by_id("user").send_keys("wangjie02@vipcode.com")   
    driver.find_element_by_id("pass").send_keys("123456")
    driver.find_element_by_id("loginBtn").click()
    sleep(1)
    vipafter = driver.get_cookies()
    driver.find_element_by_xpath('//*[@id="updateSystem"]/div/div/div[1]/button').click()
    sleep(2)
    driver.find_element_by_link_text("教学管理").click()
    sleep(1)
    driver.find_element_by_link_text("老师课表").click()
    sleep(1)
    table_all=[]
    # 点击 下一天 或者 上一天
    if date>0:
        for x in range(date):
            driver.find_element_by_xpath("/html/body/div[@class='main-container container-fluid']/div[@class='page-container']/div[@class='page-content']/div[@class='page-body']/div[@class='row']/div[@class='col-lg-12 col-sm-12 col-xs-12']/div[@class='widget']/div[@class='widget-body']/div[@class='bootstrap-table']/div[@class='fixed-table-toolbar']/div[@class='bars pull-left']/div[@id='toolbar']/button[@id='nextDay']").click()
            sleep(1)

            elements_all=driver.find_element_by_xpath("/html/body/div[@class='main-container container-fluid']/div[@class='page-container']/div[@class='page-content']/div[@class='page-body']/div[@class='row']/div[@class='col-lg-12 col-sm-12 col-xs-12']/div[@class='widget']/div[@class='widget-body']/div[@class='bootstrap-table']/div[@class='fixed-table-container']")
            table=elements_all.text.split("\n")
            #初步页面信息收集    
            date=table[2]        #日期
            table=table[11:-1]   #删除无用信息
            #sleep(1)
            # 信息列表数据清洗
            table_new=[[date]]   # 有课时间段 学生信息：时间-第几课时-课时详情-
            for x in range(len(table)):
                if len(table[x])==11:
                    k=table[x:(x+4)]                  
                    if "P1" in k[2]:
                        k[1]="P1_"+k[1]
                    elif "P2" in k[2]:
                        k[1]="P2_"+k[1]     
                    k[3]=getStudent(k[3])
                    k.append(k[3]+"明天的课程是在"+k[0][:5:]+"开始哦！可以让孩子提前温习一下。")
                    del k[2]
                    table_new.append(k)
            table_all.append(table_new) 
            print(x)       
    elif date<0:
        for x in range(-date):
            driver.find_element_by_xpath("/html/body/div[@class='main-container container-fluid']/div[@class='page-container']/div[@class='page-content']/div[@class='page-body']/div[@class='row']/div[@class='col-lg-12 col-sm-12 col-xs-12']/div[@class='widget']/div[@class='widget-body']/div[@class='bootstrap-table']/div[@class='fixed-table-toolbar']/div[@class='bars pull-left']/div[@id='toolbar']/button[@id='prevDay']").click()
            sleep(1)

            elements_all=driver.find_element_by_xpath("/html/body/div[@class='main-container container-fluid']/div[@class='page-container']/div[@class='page-content']/div[@class='page-body']/div[@class='row']/div[@class='col-lg-12 col-sm-12 col-xs-12']/div[@class='widget']/div[@class='widget-body']/div[@class='bootstrap-table']/div[@class='fixed-table-container']")
            table=elements_all.text.split("\n")
            #初步页面信息收集    
            date=table[2]        #日期
            table=table[11:-1]   #删除无用信息
            #sleep(1)
            # 信息列表数据清洗
            table_new=[[date]]   # 有课时间段 学生信息：时间-第几课时-课时详情-
            for x in range(len(table)):
                if len(table[x])==11:
                    k=table[x:(x+4)]                  
                    if "P1" in k[2]:
                        k[1]="P1_"+k[1]
                    elif "P2" in k[2]:
                        k[1]="P2_"+k[1]    
                    k[3]=getStudent(k[3])
                    #k.append(k[3]+"今天的课程是在"+k[0][:5:]+"开始哦！可以让孩子提前温习一下。")
                    del k[2]
                    table_new.append(k)
            table_all.append(table_new)
            print(x)  
    else:
        elements_all=driver.find_element_by_xpath("/html/body/div[@class='main-container container-fluid']/div[@class='page-container']/div[@class='page-content']/div[@class='page-body']/div[@class='row']/div[@class='col-lg-12 col-sm-12 col-xs-12']/div[@class='widget']/div[@class='widget-body']/div[@class='bootstrap-table']/div[@class='fixed-table-container']")
        table=elements_all.text.split("\n")
        #初步页面信息收集    
        date=table[2]        #日期
        table=table[11:-1]   #删除无用信息
        #sleep(1)
        # 信息列表数据清洗
        table_new=[[date]]   # 有课时间段 学生信息：时间-第几课时-课时详情-
        for x in range(len(table)):
            if len(table[x])==11:
                    k=table[x:(x+4)]                  
                    if "P1" in k[2]:
                        k[1]="P1_"+k[1]
                    elif "P2" in k[2]:
                        k[1]="P2_"+k[1]    
                    k[3]=getStudent(k[3])
                    k.append(k[3]+"今天的课程是在"+k[0][:5:]+"开始哦！可以让孩子提前温习一下。这个PDF文件是孩子今天的课程导读，可以提前预习。")
                    del k[2]
                    table_new.append(k)               
        table_all.append(table_new)                            
    return table_all      
def listExcel(list_in):
    wb = load_workbook(filename="temp.xlsx")
    ws=wb.active
    sheets = wb.get_sheet_names()
    sheet_first = sheets[0]
    for x in list_in:
        for i in x:
            ws.append(i)
    wb.save(filename="tablelist.xlsx")        
def getStudent(studentNum):
    student={
        "P2RC220181202-3":"子熙",
        "P1RC120190106-1":"建宇",
        "P1RC120181228-1":"Simon",
        "P1RC220181229-3":"小秋",
        "P1RC120181221-3":"诗钧",
        "P2RC120190110-1":"昕宸",
        "P1RC120190120-1":"大牛",
        "P1RC220190224-2":"梓涵",
        "P1RC120190225-3":"靖航",
        "P1RC120190223-5":"美鹤",
        "P1RC120190223-6":"涵硕",
        "P1RC220190223-6":"荣枭",
        "P1RC120190226-2":"子墨",
        "P1RC120190223-3":"莫涵",
        "P1RC120190303-6":"Ike",
        "P1RC120190310-15":"小鑫",
        "P1RC120190310-16":"萱萱",
        "P1RC420180630-3":"延昊",
        "P1RC220181207-1":"航锐、松松",
        "P1RC220181217-1":"裕宸、家乐",
        "P2RC220190108-1":"小哲、民松",
        "P1RC220190126-3":"欣怡、阳阳",
        "P1RC220190125-2":"见吾、小毛",
        "P1RC120190401-4":"小展",
        "P1RC120190223-8":"宏峻"
    } 
    name=studentNum 
    for key in student:
        if key==studentNum:
            name=student[key]
    return name
def getTableStr():
    wb = load_workbook("tablelist.xlsx")
    ws=wb.active
    i=3
    msg=[]
    s=ws["A2"].value+"\n\n"
    while True:
        word=ws["B"+str(i)].value
        if ws["B"+str(i)].value == None:
            break    
        word1=ws["C"+str(i)].value
        try:
            if word[5]=="课":
                msg.append([word[:2],"0"+word[4],word1]) 
            else:    
                msg.append([word[:2],word[4:6],word1])
            word2=ws["A"+str(i)].value
            word3=ws["D"+str(i)].value
            s += word2+" "+word+" "+word1+"\n\n"+ word3+"\n\n\n"
            i+=1 
        except IndexError:
            i+=1     


    return [msg,s] 
def getStr(tli,d):
    s=""
    for x in tli[d]:
        h=""
        for y in x:
            h+=y+" "
        s+=h+"\n"    
    return s
if __name__ == "__main__":
    # 获得课表数据
    #table=find_taday_table(11)
    # 将课表数据写入表格中
    #listExcel(table)
    # 获取课表的字符串格式

    h=getTableStr()[1]
    print(h)
    f=open("text.txt","w")
    f.write(h)
    f.close()
